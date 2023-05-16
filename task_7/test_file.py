import os
import csv
import time
import xlrd
import os.path
import requests

from pypdf import PdfReader
from selenium import webdriver
from openpyxl import load_workbook
from selene.support.shared import browser
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# TODO оформить в тест, добавить ассерты и использовать универсальный путь
PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
RESOURCES_PATH = os.path.join(PROJECT_ROOT_PATH, 'resources')


def test_csv():
    csv_path = os.path.join(RESOURCES_PATH, 'example.csv')
    with open(csv_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_path) as csvfile:
        csvreader = csv.reader(csvfile)
        arr = []
        for row in csvreader:
            arr.append(row)

    assert arr[0] == ['Anna', 'Pavel', 'Peter']
    assert arr[1] == ['Alex', 'Serj', 'Yana']


def test_download_file_with_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": RESOURCES_PATH,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=options
    )
    browser.config.driver = driver
    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()

    time.sleep(10)

    download_file = os.path.join(RESOURCES_PATH, 'pytest-main.zip')
    assert os.path.exists(download_file)


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'

    r = requests.get(url)
    tmp = os.path.join(RESOURCES_PATH, 'selenium_logo.png', )
    with open(tmp, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(tmp)

    assert size == 30803


def test_pdf():
    pdf_path = os.path.join(RESOURCES_PATH, "docs-pytest-org-en-latest.pdf")
    with open(pdf_path, 'rb') as file:
        reader = PdfReader(file)
        number_of_pages = len(reader.pages)
        page = reader.pages[0]
        text = page.extract_text()
        print(page)
        print(number_of_pages)
        print(text)
        assert number_of_pages == 412
        assert text == 'pytest Documentation\n' \
                       'Release 0.1\n' \
                       'holger krekel, trainer and consultant, https://merlinux.eu/\n' \
                       'Jul 14, 2022'


def test_xls():
    file = os.path.join(RESOURCES_PATH, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(file)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=9, colx=1) == 'Vincenza'


def test_xlsx():
    xlsx_file = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    assert sheet.cell(row=2, column=5).value == 'United States'
